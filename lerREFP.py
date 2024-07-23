import PySimpleGUI as sg
import openpyxl
from openpyxl import load_workbook

janela_principal = [  
            [sg.Text("Qual seria o nome do arquivo origem? (coloque o caminho)")],
            [sg.InputText(key="arquivo_Origem")],
            [sg.Text("Qual seria o nome do arquivo destino? (coloque o caminho)")],
            [sg.InputText(key="arquivo_Destino")],
            [sg.Button("P&D")], [sg.Button("PE")], [sg.Button("Cancelar")] 
            ]

window_principal = sg.Window("Dados", janela_principal)

while True:

    event_principal, values_principal = window_principal.read()

    window_principal.close()

    if event_principal == sg.WIN_CLOSED or event_principal == 'Cancelar':
        break

    if event_principal == 'P&D':
        print("Chegou no P&D")
        
        arquivoDeOrigem = values_principal["arquivo_Origem"]
        arquivoDeDestino = values_principal["arquivo_Destino"]

        janela_XML = [
                [sg.Text("Quem está descrito no XML?")],
                [sg.InputText(key="descrito_XML")],
                [sg.Button("Proseguir")], [sg.Button("Cancelar")]
            ]
        
        window_XML = sg.Window("Descritos no XML", janela_XML)
        
        while True:

            event_XML, values_XML = window_XML.read()
            
            window_XML.close()

            if event_XML == "Cancelar":
            
                janela_problema = [
                    [sg.Text("Tem certeza? Isso irá fechar o programa!")],
                    [sg.Button("Sim")], [sg.Button("Não")]
                ]

                window_problema = sg.Window("Está correto?", janela_problema)
        
                while True:

                    event_problema, values_problema = window_problema.read()
                    
                    window_problema.close()

                    if event_problema == "Não":
                        ##colocar uma função de pesquisa para P&D
                        ##colocar uma função para se não houver ninguém descrito
                        ##colocar janela nova, mas sem poder cancelar
                        break

                    if event_problema == "Sim":
                        break

            if event_XML == "Proseguir":
                print("Colocar função para quando tem e quando não tem ninguém descrito")
                ##colocar uma função de pesquisa para P&D
                ##colocar uma função para se não houver ninguém descrito
            break

    if event_principal == 'PE':
        print("Chegou no PE")
        ##Usar alguma notificação de que a linha de preenchimento está em branco as linhas preenchidas
        
        arquivoDeOrigem = values_principal["arquivo_Origem"]
        arquivoDeDestino = values_principal["arquivo_Destino"]

        for nomeSemAspas in range(len(arquivoDeOrigem)): ## Tirar aspas da cópia de local
            arquivoDeOrigem = arquivoDeOrigem.replace('"', "")

####################################################################
        ##Colocar uma função de pesquisa para PE

        workbook = openpyxl.load_workbook(filename=arquivoDeOrigem)

        janela_escolha = [
                    [sg.Text("Qual seria a aba?")],
                    [sg.Text(workbook.sheetnames)],
                    [sg.InputText(key='escolha_Aba')],
                    [sg.Button('Usar esta aba!')], [sg.Button('Cancelar')]
                ]
        window_escolha = sg.Window("Escolha", janela_escolha)

        while True:

            event_escolha, values_escolha = window_escolha.read()

            window_escolha.close()

            if event_escolha == sg.WIN_CLOSED or event_escolha == 'Cancelar':
                break

            if event_escolha == 'Usar esta aba!':
                
                abaEscolhida = values_escolha['escolha_Aba']

                sheet = workbook[abaEscolhida] 

                for row in range(1, 5001):
                    for column in range(1, 31):
                        celula = sheet.cell(row, column) ##Escolhe célula específica para modificar
                        celula.value = 'Hello World!' ##Adiciona um valor para a célula

                workbook.save(arquivoDeOrigem) ##Salva planiha IMPORTANTE

#####################################################################
        break